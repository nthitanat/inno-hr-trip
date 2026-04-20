"""
pipeline.py — Main orchestrator
Runs all 3 steps in sequence:
  Step 1: BuildTripFlags   → rawdata_flagged + trips sheets
  Step 2: BuildTripBehavior → trip_behavior sheet
  Step 3: BuildTripCleaned  → trip_cleaned sheet + summary stats

All config is read from environment variables (loaded from .env via docker-compose).
"""

import logging
import os
import sys
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

from step1_flags import build_trip_flags
from step2_behavior import build_trip_behavior
from step3_cleaned import build_trip_cleaned

# ---------------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------------
LOG_LEVEL = os.getenv("LOG_LEVEL", "INFO").upper()
logging.basicConfig(
    level=getattr(logging, LOG_LEVEL, logging.INFO),
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
log = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Load config from environment (all set in .env → docker-compose → container)
# ---------------------------------------------------------------------------
def load_config() -> dict:
    return {
        # Thresholds
        "SAME_LOC_M":      float(os.getenv("SAME_LOC_M",      30)),
        "BORDERLINE_M":    float(os.getenv("BORDERLINE_M",     50)),
        "SHORT_STOP_MIN":  float(os.getenv("SHORT_STOP_MIN",    5)),
        "NIGHT_HR_START":  int(os.getenv("NIGHT_HR_START",      0)),
        "NIGHT_HR_END":    int(os.getenv("NIGHT_HR_END",         5)),
        "NEXTDAY_MIN":     float(os.getenv("NEXTDAY_MIN",      720)),
        # Column indices — VBA 1-based; Python subtracts 1 inside step functions
        "COL_BRANCH":   int(os.getenv("COL_BRANCH",   2)),
        "COL_PERSON":   int(os.getenv("COL_PERSON",   7)),
        "COL_TITLE":    int(os.getenv("COL_TITLE",    8)),
        "COL_FNAME":    int(os.getenv("COL_FNAME",    9)),
        "COL_LNAME":    int(os.getenv("COL_LNAME",   10)),
        "COL_DATETIME": int(os.getenv("COL_DATETIME",13)),
        "COL_LAT":      int(os.getenv("COL_LAT",     18)),
        "COL_LNG":      int(os.getenv("COL_LNG",     19)),
        "COL_GPS_ACC":  int(os.getenv("COL_GPS_ACC", 20)),
        "COL_LOCNAME":  int(os.getenv("COL_LOCNAME", 21)),
    }


# ---------------------------------------------------------------------------
# Read input file
# ---------------------------------------------------------------------------
def read_input(input_path: str, sheet_name: str) -> pd.DataFrame:
    p = Path(input_path)
    if not p.exists():
        raise FileNotFoundError(f"Input file not found: {p.resolve()}")

    log.info(f"Reading input: {p.name}  sheet='{sheet_name}'")
    ext = p.suffix.lower()

    if ext == ".csv":
        df = pd.read_csv(input_path, encoding="utf-8-sig", header=0)
    elif ext in (".xlsx", ".xlsm", ".xls"):
        df = pd.read_excel(input_path, sheet_name=sheet_name, header=0)
    else:
        raise ValueError(f"Unsupported file type: {ext}")

    log.info(f"  Loaded {len(df)} rows × {len(df.columns)} columns")
    return df


# ---------------------------------------------------------------------------
# Write output Excel workbook (all 5 sheets)
# ---------------------------------------------------------------------------
HEADER_FILL  = PatternFill("solid", fgColor="1F3864")   # dark navy
HEADER_FONT  = Font(bold=True, color="FFFFFF")
SUMMARY_FILL = PatternFill("solid", fgColor="D6E4F0")   # light blue
ALT_FILL     = PatternFill("solid", fgColor="F0F4F8")   # light grey stripe


def _write_df_to_sheet(ws, df: pd.DataFrame):
    """Write DataFrame to openpyxl worksheet with header styling."""
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
                cell.alignment = Alignment(horizontal="center")


def _write_summary_to_sheet(ws, summary: dict, start_row: int):
    """Append summary stats below the data in trip_cleaned sheet."""
    srow = start_row + 2

    # Section header
    cell = ws.cell(row=srow, column=1, value="SUMMARY — Clean Trips for Budget Calculation")
    cell.font = Font(bold=True, size=11)
    cell.fill = SUMMARY_FILL
    srow += 1

    # Column labels
    ws.cell(row=srow, column=1, value="Metric").font = Font(bold=True, color="FFFFFF")
    ws.cell(row=srow, column=1).fill = HEADER_FILL
    ws.cell(row=srow, column=2, value="Value").font = Font(bold=True, color="FFFFFF")
    ws.cell(row=srow, column=2).fill = HEADER_FILL
    srow += 1

    rows = [
        ("Total trips (before cleaning)",    summary["total_before"]),
        ("Removed: Forget Checkout",         summary["removed_forget"]),
        ("Removed: New Day Work",            summary["removed_newday"]),
        ("Total removed",                    summary["total_removed"]),
        ("Clean trips (for budget)",         summary["total_kept"]),
        ("Total distance (m)",               round(summary["total_distance_m"], 1)),
        ("Total distance (km)",              round(summary["total_distance_km"], 3)),
        ("Average distance per trip (m)",    round(summary["avg_distance_m"], 1)),
        ("Average duration per trip (min)",  round(summary["avg_duration_min"], 1)),
    ]
    for label, value in rows:
        ws.cell(row=srow, column=1, value=label)
        ws.cell(row=srow, column=2, value=value)
        srow += 1

    # Per-person subtotals
    srow += 1
    cell = ws.cell(row=srow, column=1, value="PER-PERSON DISTANCE SUMMARY")
    cell.font = Font(bold=True, size=11)
    cell.fill = SUMMARY_FILL
    srow += 1

    pp_headers = ["Person_ID", "Full_Name", "Branch",
                  "Trip_Count", "Total_Distance_m", "Total_Distance_km", "Avg_Distance_m"]
    for c_idx, h in enumerate(pp_headers, 1):
        cell = ws.cell(row=srow, column=c_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    srow += 1

    for i, pp in enumerate(summary.get("per_person", [])):
        fill = ALT_FILL if i % 2 == 1 else None
        vals = [pp["person_id"], pp["full_name"], pp["branch"],
                pp["trip_count"], round(pp["total_dist_m"], 1),
                round(pp["total_dist_km"], 3), round(pp["avg_dist_m"], 1)]
        for c_idx, v in enumerate(vals, 1):
            cell = ws.cell(row=srow, column=c_idx, value=v)
            if fill:
                cell.fill = fill
        srow += 1


def write_output(output_path: str,
                 rawdata_df: pd.DataFrame,
                 flagged_df: pd.DataFrame,
                 trips_df: pd.DataFrame,
                 behavior_df: pd.DataFrame,
                 cleaned_df: pd.DataFrame,
                 summary: dict):
    """Write all 5 sheets into one Excel workbook."""
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)   # remove default empty sheet

    sheets = [
        ("rawdata",          rawdata_df),
        ("rawdata_flagged",  flagged_df),
        ("trips",            trips_df),
        ("trip_behavior",    behavior_df),
        ("trip_cleaned",     cleaned_df),
    ]

    for sheet_name, df in sheets:
        ws = wb.create_sheet(title=sheet_name)
        _write_df_to_sheet(ws, df)
        log.info(f"  Sheet '{sheet_name}': {len(df)} rows written")

    # Append summary stats below trip_cleaned data
    ws_cleaned = wb["trip_cleaned"]
    _write_summary_to_sheet(ws_cleaned, summary, start_row=len(cleaned_df) + 1)

    wb.save(output_path)
    log.info(f"Output saved: {output_path}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    log.info("=" * 60)
    log.info("  Trip Pipeline: rawdata → flagged → trips → behavior → cleaned")
    log.info("=" * 60)

    config = load_config()

    input_file  = os.getenv("INPUT_FILE",  "input/rawdata.xlsx")
    output_file = os.getenv("OUTPUT_FILE", "output/result.xlsx")
    sheet_name  = os.getenv("RAWDATA_SHEET", "rawdata")

    # ── Step 0: Read input ──────────────────────────────────────────
    rawdata_df = read_input(input_file, sheet_name)

    # ── Step 1: BuildTripFlags ──────────────────────────────────────
    log.info("[Step 1] BuildTripFlags — flagging + trip pairing")
    flagged_df, trips_df = build_trip_flags(rawdata_df, config)
    log.info(f"  rawdata_flagged: {len(flagged_df)} rows")
    log.info(f"  trips:           {len(trips_df)} rows")

    # ── Step 2: BuildTripBehavior ───────────────────────────────────
    log.info("[Step 2] BuildTripBehavior — behavior flags")
    behavior_df = build_trip_behavior(trips_df, config)
    log.info(f"  trip_behavior:   {len(behavior_df)} rows")

    # ── Step 3: BuildTripCleaned ────────────────────────────────────
    log.info("[Step 3] BuildTripCleaned — filtering + summary stats")
    cleaned_df, summary = build_trip_cleaned(behavior_df, config)
    log.info(f"  trip_cleaned:    {len(cleaned_df)} rows (clean trips)")
    log.info(f"  Removed:         {summary['total_removed']} rows")
    log.info(f"  Total distance:  {summary['total_distance_km']:.2f} km")

    # ── Write output ────────────────────────────────────────────────
    log.info("[Output] Writing Excel workbook...")
    write_output(output_file, rawdata_df, flagged_df,
                 trips_df, behavior_df, cleaned_df, summary)

    log.info("=" * 60)
    log.info("  Pipeline complete!")
    log.info(f"  Clean trips : {summary['total_kept']}")
    log.info(f"  Total km    : {summary['total_distance_km']:.2f}")
    log.info(f"  Output      : {output_file}")
    log.info("=" * 60)


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        log.error(f"Pipeline failed: {e}", exc_info=True)
        sys.exit(1)
